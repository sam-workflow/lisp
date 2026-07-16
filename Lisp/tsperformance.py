'''
Created on Nov 22, 2016

@author: thomas
'''
import openpyxl
import glob
import os
import re
import sys


def main(strDirectory):
    market = ''
    comboPerformance = []
    for filename in glob.iglob(os.environ['HOME'] + '/mk-data/111-Dropbox/' + strDirectory[0] + '/*/*.xlsx'):
        if market != regCheck('.*\/([A-Z]*)\/.*\.xlsx', filename)[0]:
            market = regCheck('.*\/([A-Z]*)\/.*\.xlsx', filename)[0]
        if regCheck('(~)', filename)[0] == '':
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
            row = 5
            while ws[''.join(['B', str(row)])].value is not None:
                if ws[''.join(['A', str(row)])].value is None:
                    date = parseDate(str(ws[''.join(['C', str(row)])].value))
                    profit = str(ws[''.join(['G', str(row)])].value)
                    comboPerformance.append(",".join([date, market, profit]))
                row += 1
        ws = None
        wb = None
    with open(os.environ['HOME'] + '/mk-data/111-Dropbox/' + strDirectory[0] + '/tsperformance.csv', 'w') as f:
        f.write('\n'.join(comboPerformance))
    print "Process Completed.."


def parseDate(strDate):
    date = regCheck('([\d\-]*)\ ', strDate)[0]
    return ''.join(date.split('-'))


def regCheck(rePattern, strItem):
    reMatch = re.compile(rePattern, re.MULTILINE | re.IGNORECASE)
    allMatches = reMatch.findall(strItem)
    if len(allMatches) > 0:
        return allMatches
    else:
        return ['']

if __name__ == '__main__':
    if len(sys.argv) > 1:
            main(sys.argv[1:])
    else:
        print "Error: No arguments!"
        print main.__doc__
